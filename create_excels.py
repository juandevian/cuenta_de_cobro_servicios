#!/usr/bin/env python3
"""
Script para crear un archivo Excel de plantilla para importación de cobros
"""
import pandas as pd
import os
from openpyxl import load_workbook  # Para formato y tabla
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def create_sample_excel(output_path=None):
    """Crea un archivo Excel de ejemplo con datos de items de facturas"""
    filename = 'ejemplo_facturas.xlsx'
    if output_path:
        full_path = os.path.join(output_path, filename)
    else:
        full_path = filename   

    # Crear datos de ejemplo
    data = {
        'id_carpeta': [5, 5, 5, 5, 5, 5, 5, 5, 5, 5],
        'id_servicio': [3, 3, 3, 3, 3, 3, 3, 3, 3, 3],
        'id_predio': ['', 'APT 102', 'APT 103', 'APT 201', 'APT 202', 'PAR 001', 'PAR 002', '', 'PAR 004', 'PAR 005'],
        'id_tercero_cliente': [10999000, 0, 0, 0, 0, 0, 0, 1053700700, 0, 0],
        'periodo_inicio_cobro': ['202609', '202609', '202609', '202609', '202609', '202609', '202609', '202609', '202609', '202609'],
        'lectura_anterior': [1000, 2500, 800, 1150, 3200, 500, 1500, 2000, 4000, 6000],
        'lectura_actual': [1150, 2750, 950, 1320, 3450, 650, 1650, 2200, 4300, 6300],
        'valor_unitario': [850, 920, 780, 850, 950, 800, 900, 870, 960, 990]
    }

    # Crear DataFrame
    df = pd.DataFrame(data)

    # Crear archivo Excel
    df.to_excel(full_path, index=False, engine='openpyxl', sheet_name='Ejemplo Items Factura')

    # Abrir con openpyxl para agregar formato
    wb = load_workbook(full_path)
    ws = wb.active

    content_font = Font(name="Arial", size=12)

    for row in ws.iter_rows():
        for cell in row:
            cell.font = content_font

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")
   

    for col in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Definir formatos de datos por columna (desde fila 2 en adelante)
    column_formats = {
        'id_carpeta': '0',  # Número entero
        'id_servicio': '0',  # Número entero
        'id_predio': '@',  # Texto
        'id_tercero_cliente': '0',  # Número entero
        'periodo_inicio_cobro': '@',  # Texto (ej. '202609')
        'lectura_anterior': '0.00',  # Número entero
        'lectura_actual': '0.00',  # Número entero
        'valor_unitario': '0'  # Decimal con 2 decimales
    }

    # Aplicar formatos a las columnas (desde fila 2)
    for col_num, col_name in enumerate(df.columns, start=1):
        format_str = column_formats.get(col_name, '@')  # Por defecto texto
        for row_num in range(2, len(df) + 2):  # Desde fila 2 hasta el final
            cell = ws.cell(row=row_num, column=col_num)
            cell.number_format = format_str

    # Ajustar anchos de columna
    column_widths = [11, 11, 22, 22, 20, 18, 18, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    # Guardar el archivo con formato y tabla
    wb.save(full_path)
    
    print(f"✓ Archivo Excel de ejemplo creado: {full_path}")
    print(f"✓ Filas: {len(df)}")
    print(f"✓ Columnas: {list(df.columns)}")
    print("✓ Fuente aplicada: Arial, tamaño 12 a todo el contenido.")
    print("\nDatos de ejemplo incluidos:")
    print(df.to_string(index=False))

    return full_path

def create_excel_import_template(output_path=None):
    """Crea una plantilla en Excel para la importación de cobros de servicios con consumo"""
    filename = 'plantilla_importacion_servicios.xlsx'
    if output_path:
        full_path = os.path.join(output_path, filename)
    else:
        full_path = filename  # Ruta por defecto

    columns = {
        'id_carpeta': [],
        'id_servicio': [],
        'id_predio': [],
        'id_tercero_cliente': [],
        'periodo_inicio_cobro': [],
        'lectura_anterior': [],
        'lectura_actual': [],
        'valor_unitario': []
    }

    df = pd.DataFrame(columns)
    df.to_excel(filename, index=False, engine='openpyxl', sheet_name='Import. Servicios')
    
    wb = load_workbook(filename)
    ws = wb.active

    # Definir fuente para todo el contenido
    content_font = Font(name="Arial", size=12)

    # Aplicar fuente a todas las celdas
    for row in ws.iter_rows():
        for cell in row:
            cell.font = content_font

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")
   

    for col in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Definir formatos de datos por columna (desde fila 2 en adelante)
    column_formats = {
        'id_carpeta': '0',  # Número entero
        'id_servicio': '0',  # Número entero
        'id_predio': '@',  # Texto
        'id_tercero_cliente': '0',  # Número entero
        'periodo_inicio_cobro': '@',  # Texto (ej. '202609')
        'lectura_anterior': '0.00',  # Número entero
        'lectura_actual': '0.00',  # Número entero
        'valor_unitario': '0'  # Decimal con 2 decimales
    }

    # Aplicar formatos a las columnas (desde fila 2)
    for col_num, col_name in enumerate(df.columns, start=1):
        format_str = column_formats.get(col_name, '@')  # Por defecto texto
        for row_num in range(2, len(df) + 2):  # Desde fila 2 hasta el final
            cell = ws.cell(row=row_num, column=col_num)
            cell.number_format = format_str

    # Ajustar anchos de columna
    column_widths = [11, 11, 22, 22, 20, 18, 18, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    # Guardar el archivo con formato y tabla
    wb.save(full_path)

    print(f"✓ Archivo de plantilla Excel creado con formato y tabla: {full_path}")
    print(f"✓ Columnas: {list(df.columns)}")
    print("✓ Formatos aplicados: números enteros para IDs/lecturas, decimales para valores, texto para predios/períodos.")
    return full_path

if __name__ == "__main__":
    create_sample_excel()